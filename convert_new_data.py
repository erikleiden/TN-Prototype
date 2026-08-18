"""
Convert refreshed TN data files from xlsx/csv to JSON for the dashboard app.

Usage:
    python convert_new_data.py --data-dir /path/to/refreshed/files

Expects these files in --data-dir (names must match the refresh delivery):
    2026.03.31_stall_crosstab.csv
    2026.03.31_stall_duration_hist_data.csv
    national_trans_clean_top5.xlsx           (sheet "data")
    occ_similarity_top5.xlsx                 (sheet "data")
    app_skill_gaps_top5.xlsx
    tennessee_posting_demand_by_occ_and_sector.xlsx
        (sheets "occ", "occ_sector", "share_growth")

Outputs JSON files into src/data/ matching the app's expected formats.
Run build_demographics.py and build_overlap.py afterwards, then `npm run build`.
"""

import argparse
import json
import csv
import os
import openpyxl

parser = argparse.ArgumentParser(description="Convert TN refresh files to app JSON.")
parser.add_argument("--data-dir", required=True,
                    help="Folder containing the refreshed CSV/XLSX delivery")
args = parser.parse_args()

DATA_DIR = args.data_dir
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "src", "data")

if not os.path.isdir(DATA_DIR):
    raise SystemExit(f"ERROR: --data-dir not found: {DATA_DIR}")


def read_xlsx_data(filename, sheet_name="data"):
    """Read data sheet from xlsx, return list of dicts."""
    path = os.path.join(DATA_DIR, filename)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def read_csv_data(filename):
    """Read CSV, return list of dicts with dynamic typing."""
    path = os.path.join(DATA_DIR, filename)
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        data = []
        for row in reader:
            typed = {}
            for k, v in row.items():
                if v == "" or v is None:
                    typed[k] = None
                else:
                    try:
                        typed[k] = int(v)
                    except ValueError:
                        try:
                            typed[k] = float(v)
                        except ValueError:
                            typed[k] = v
            data.append(typed)
    return data


def write_json(data, filename):
    path = os.path.join(OUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"  Wrote {path} ({len(data) if isinstance(data, list) else 'dict'} entries)")


# ============================================================================
# 1. Stall Crosstab → cross_tabulated_data.json
# ============================================================================
print("Converting stall crosstab...")
stall_data = read_csv_data("2026.03.31_stall_crosstab.csv")
write_json(stall_data, "cross_tabulated_data.json")


# ============================================================================
# 2. Stall Duration Histogram → stall_duration.json
# ============================================================================
print("Converting stall duration histogram...")
duration_data = read_csv_data("2026.03.31_stall_duration_hist_data.csv")
write_json(duration_data, "stall_duration.json")


# ============================================================================
# 3. National Transitions → national_transitions.json
#    Rename _origin/_destination → _SOURCE/_TARGET to match app's PathwayRow
# ============================================================================
print("Converting national transitions...")
trans_raw = read_xlsx_data("national_trans_clean_top5.xlsx", "data")
trans_out = []
for row in trans_raw:
    mapped = {}
    for k, v in row.items():
        # Rename _origin → _SOURCE, _destination → _TARGET
        new_k = k.replace("_origin", "_SOURCE").replace("_destination", "_TARGET")
        # Also rename the 5-year column names
        new_k = new_k.replace("internal_promotion_rate_5_year_", "internal_promotion_rate_5_")
        new_k = new_k.replace("external_promotion_rate_5_year_", "external_promotion_rate_5_")
        mapped[new_k] = v
    trans_out.append(mapped)
write_json(trans_out, "national_transitions.json")


# ============================================================================
# 4. Occupation Similarity → occ_similarity.json
#    Already uses _SOURCE/_TARGET naming
# ============================================================================
print("Converting occupation similarity...")
sim_raw = read_xlsx_data("occ_similarity_top5.xlsx", "data")
sim_out = []
for row in sim_raw:
    mapped = {}
    for k, v in row.items():
        mapped[k] = v
    sim_out.append(mapped)
write_json(sim_out, "occ_similarity.json")


# ============================================================================
# 5. Skill Gaps → skill_gaps_top5.json
#    (top20 was superseded by cross_pathway_skills.json and is no longer built)
# ============================================================================
print("Converting skill gaps (top5)...")
sg5 = read_xlsx_data("app_skill_gaps_top5.xlsx")
write_json(sg5, "skill_gaps_top5.json")


# ============================================================================
# 6. Posting Demand → posting_demand.json
#    Multi-key dict: { occ, occ_sector, share_growth }
# ============================================================================
print("Converting posting demand...")
demand_occ = read_xlsx_data("tennessee_posting_demand_by_occ_and_sector.xlsx", "occ")
demand_occ_sector = read_xlsx_data("tennessee_posting_demand_by_occ_and_sector.xlsx", "occ_sector")
demand_growth = read_xlsx_data("tennessee_posting_demand_by_occ_and_sector.xlsx", "share_growth")

posting_demand = {
    "occ": demand_occ,
    "occ_sector": demand_occ_sector,
    "share_growth": demand_growth,
}
write_json(posting_demand, "posting_demand.json")


print("\nDone! All JSON files updated in src/data/")
